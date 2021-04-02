from time import strptime
from sqlalchemy.sql.sqltypes import Time
#from app.views import get_user
from os import abort, name
from app.models import Project, Activity, Customer, Order, Cuos, Tasks, Timesheet
from app import db
import openpyxl
from config import UPLOAD_FOLDER
from datetime import timedelta, datetime
from openpyxl import load_workbook

def upload_project():
        session = db.session
        prj = open('xls/progetti.csv', encoding='utf-8-sig')
        for line in prj:
                row = Project(project=line)
                row.created_by_fk = '1'
                row.changed_by_fk = '1'
                session.add(row)
        session.commit()

def upload_activity():
        session = db.session
        prj = open('xls/activities.csv', encoding='utf-8-sig')
        for line in prj:
                row = Activity(activity=line)
                row.created_by_fk = '1'
                row.changed_by_fk = '1'
                session.add(row)
        
        session.commit()

def upload_history():
        session = db.session
        projects = dict([(x.project, x.id) for x in session.query(Project).all()])
        activities = dict([(x.activity, x.id) for x in session.query(Activity).all()])
        act = open('xls/history.csv')

def upload_pm_items():
    session = db.session
    wb = load_workbook('xls/PM_items.xlsx')
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        customer = row[0].value
        order = row[1].value 
        project = row[2].value
        cuo = row[3].value
        activity = row[4].value
        activity_type = row[5].value

from flask import g      
def get_user_id():
    return g.user.id

def get_user():
    return g.user


from zip_helper import read_zip
from flask_appbuilder.security.sqla.manager import User
from flask import flash 
'''
def check_tbd(item, deliverable):
    # verifica l'esistenza di billables/items con transmittal TBD (to be defined) e ne aggiorna il transmittal
    session = db.session
    billables = session.query(Billitem).filter(Billitem.deliverable == 'TBD', Billitem.item == item).all()
    for bill in billables:
        bill.deliverable = deliverable
    session.commit()

def update_billable(items):
    session = db.session
    print('update billable started')
    print('file path', UPLOAD_FOLDER,'file', items.billable)
    file_list = []
    if str(items.billable).split('.')[1] == 'zip':
        zip_items = read_zip(UPLOAD_FOLDER + items.billable)
        file_list = zip_items[0]
    
    elif str(items.billable).split('_sep_')[1] == 'tr.xlsx':
        try:
            print('open TR.XLSX file')
            ts = openpyxl.load_workbook(UPLOAD_FOLDER + items.billable)
            print('TR.XLSX file OPEND')
            bill_ws = bill_file.active
        
            for row in bill_ws.iter_rows(min_row=4):
                    print(row[0].value)
                    if row[0].value and row[1].value:
                        if row[2].value is None:
                            billitem = Billitem(
                                            tasks_id=items.id,
                                            #timesheet_id=items.timesheet_id,
                                            deliverable = row[0].value,
                                            doc_quantity = row[1].value,
                                            item=row[2].value,
                                            time=row[3].value,
                                            comments=row[4].value
                                            )

                            #session.add(billitem)
                        else:
                            billitem = Billitem(
                                            tasks_id=items.id,
                                            #timesheet_id=items.timesheet_id,
                                            deliverable = row[0].value,
                                            doc_quantity = row[1].value,
                                            item=row[2].value,
                                            time=row[3].value,
                                            comments=row[4].value
                                            )
                        
                        # add only if the deliverable is unique
                        billables = session.query(Billitem).filter(Billitem.item == billitem.item, Billitem.deliverable == 'TBD').all()
                        for bill in billables:
                            bill.deliverable = billitem.deliverable
                        session.add(billitem)
                        print('added:',billitem)
                    #session.commit()
        except:
            print('EXCEPTION')
            pass
    else:
        file_list.append(items.billable)
    try:
        for file in file_list:
            
            if file[:2] != '__' and '.' in file and file != 'tr.xlsx':
                print(file)
                billitem = Billitem(
                        tasks_id=items.id,
                        #timesheet_id=items.timesheet_id,
                        deliverable = 'ND',
                        item=file,
                        time='0.25',
                        doc_quantity = 1
                        )
                session.add(billitem)
                print('added:',billitem)
    except:
        print('Errore lista file')
        pass


    session.commit()
    print('session committed **** ----- ////') 
'''

def tasks_update(self,item):
    session = db.session
    print('update tasks started')
    session.query(Tasks).delete()
    print('file path', UPLOAD_FOLDER,'file', item.file)
    file_list = []
    if str(item.file).split('.')[1] != 'xlsx':
        flash('Your TimeSheet in not in XLSX format.')
        abort(302, 'Invalid Timesheet Format')
                                                
    else:
        print('*********')
        print('*******',(UPLOAD_FOLDER + item.file))
        filepath = UPLOAD_FOLDER + item.file
        
        ts_file = openpyxl.load_workbook(filepath, read_only=True, data_only=True )
        
        print('Ts File file OPEND')
        ts_ws = ts_file.active
        row_count = 1
        
        for row in ts_ws.iter_rows(min_row=2):
            try:
                row_count += 1
                print(row_count)
                print(row[0].value)
                print(row[1].value)
                print(row[2].value)
                
                if row[0].value and row[1].value and row[2].value:
                    print('here we are ............... -1 for user id:',get_user_id())
                    # If Timesheet for this user is not there create a new one.
                    try:
                        print('try query timesheet')
                        this_ts = session.query(Timesheet).filter(Timesheet.created_by_fk== get_user_id(), Timesheet.date == str(row[0].value) ).first()
                    except:
                        print('query ts wrong ')
                    if this_ts is None:
                        print('this_ts is none')
                        this_ts = Timesheet(
                            date = row[0].value,
                            status_id = 1
                        )
                    print('Check and set the Activity/Product')
                    this_activity = session.query(Activity).filter(Activity.name == row[3].value).first()
                    if this_activity is None:
                        print('this_activity is none')
                        #flash('Activity "'+ row[3].value + '" not found.'+' Row:'+row_count, category='warning' )
                        #return abort(302, 'Product/Activity not Found')
                        ### FAKE TEST ACCEPT NEW ACTIVITY
                        this_activity = Activity(
                            name = row[3].value,
                            cuos_id = 1,
                            activity_type_id = 1,
                            status_id = 1,
                            bill_file_required = True
                        )
                    print('here we are.........2',str(row[0].value)[:10])
                    print(str(row[0].value)[:10] + " " + str(row[1].value))
                    print(row[1].value)
                    taskitem = Tasks(
                                    timesheet = this_ts,
                                    activity = this_activity,
                                    date_from = str(row[0].value)[:10] + " " + str(row[1].value),
                                    date_to = str(row[0].value)[:10] + " " + str(row[2].value),
                                    sal_item = row[4].value,
                                    ref_item= row[5].value,
                                    doc_quantity = row[6].value,
                                    comments=row[9].value
                                    )
                    print('here we are.........3')
                    session.add(taskitem)
                    session.commit()
                    print('here we are.........4')
            except Exception as e:
                print('something goes wrong :(')
                print(e)
    #session.commit()
    print('session committed **** ----- ////')
    return flash('Timesheet updated with ' + str(row_count) + ' tasks.') 


                        